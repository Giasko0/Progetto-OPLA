from flask import Blueprint, request, jsonify, session, send_file
from datetime import datetime, timedelta, date, time
import io
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from db import get_db_connection, release_connection
from auth import require_auth
from exams import controlla_vincoli, inserisci_esami

import_bp = Blueprint('import_bp', __name__)

@import_bp.route('/api/get-exam-template')
@require_auth
def get_exam_template():
    """Genera template Excel per inserimento esami."""
    username = request.args.get('docente')
    anno = request.args.get('anno')
    
    if not username:
        return jsonify({"error": "Username mancante"}), 400
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Query per insegnamenti del docente con nomi CdS
        query = """
            SELECT DISTINCT i.codice, i.titolo, c.nome_corso, c.codice as cds_codice
            FROM insegnamenti i
            JOIN insegnamento_docente id ON i.id = id.insegnamento
            JOIN insegnamenti_cds ic ON i.id = ic.insegnamento
            JOIN cds c ON ic.cds = c.codice AND ic.anno_accademico = c.anno_accademico
            WHERE id.docente = %s
        """
        
        params = [username]
        if anno:
            query += " AND id.annoaccademico = %s"
            params.append(anno)
            
        query += " ORDER BY c.nome_corso, i.titolo"
        cursor.execute(query, params)
        insegnamenti = cursor.fetchall()
        
        # Recupera aule
        cursor.execute("SELECT nome FROM aule WHERE nome != 'Studio docente DMI' ORDER BY nome")
        aule = [row[0] for row in cursor.fetchall()]
        aule.append("Studio docente DMI")  # Aggiungi in fondo
        
        cursor.close()
        release_connection(conn)
        
        if not insegnamenti:
            return jsonify({"error": "Nessun insegnamento trovato"}), 404
        
        # Crea Excel con struttura semplificata
        wb = Workbook()
        ws = wb.active
        ws.title = "Esami"
        
        # Headers aggiornati con formati specificati
        headers = [
            "CdS", 
            "Insegnamento (Descrizione)", 
            "Codice Insegnamento", # Colonna tecnica nascosta
            "Codice CdS", # Colonna tecnica nascosta
            "Apertura Appelli",
            "Data (DD-MM-YYYY)", 
            "Ora (HH:MM)", 
            "Durata (minuti)",
            "Aula", 
            "Inizio Iscrizione (DD-MM-YYYY)", 
            "Fine Iscrizione (DD-MM-YYYY)",
            "Verbalizzazione", 
            "Tipo Esame", 
            "Note"
        ]
        
        # Imposta headers con stile
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
        # Larghezza colonne
        column_widths = [30, 45, 0, 0, 18, 21, 14, 16, 20, 25, 25, 32, 15, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Righe con valori precompilati
        num_rows_to_add = len(insegnamenti)
        
        for row_idx in range(2, num_rows_to_add + 2):
            ins_data = insegnamenti[row_idx - 2] if row_idx - 2 < len(insegnamenti) else None
            
            if ins_data:
                codice_ins, titolo_ins, nome_cds, cds_codice = ins_data
                
                # CdS Nome con codice tra parentesi (precompilato)
                cds_display = f"{nome_cds} ({cds_codice})"
                cell = ws.cell(row_idx, 1, cds_display)
                cell.font = Font(name="Arial")
                cell.alignment = Alignment(horizontal="left")
                
                # Insegnamento (precompilato)
                cell = ws.cell(row_idx, 2, titolo_ins)
                cell.font = Font(name="Arial")
                cell.alignment = Alignment(horizontal="left")
                
                # Colonne nascoste (codici tecnici) subito dopo Insegnamento
                ws.cell(row_idx, 3, codice_ins)  # Codice insegnamento (nascosto)
                ws.cell(row_idx, 4, cds_codice)  # Codice CdS (nascosto)
                
                # Apertura Appelli (precompilato con "Sì")
                cell = ws.cell(row_idx, 5, "Sì")
                cell.font = Font(name="Arial")
                cell.alignment = Alignment(horizontal="left")
                
                # Verbalizzazione (precompilato)
                cell = ws.cell(row_idx, 12, "Prova finale con pubblicazione")
                cell.font = Font(name="Arial")
                cell.alignment = Alignment(horizontal="left")
            
            # Imposta stile per le celle vuote da compilare
            empty_columns = [6, 7, 8, 9, 10, 11, 13, 14]  # Data, Ora, Durata, Aula, Date iscrizione, Tipo Esame, Note
            for col in empty_columns:
                cell = ws.cell(row_idx, col)
                cell.font = Font(name="Arial")
                cell.alignment = Alignment(horizontal="left")
                cell.fill = PatternFill(fill_type=None)  # Sfondo trasparente
                
                # Imposta formati specifici per tipo di dato
                if col == 4 or col == 8 or col == 9:  # Colonne Data (formato italiano)
                    cell.number_format = 'DD-MM-YYYY'
                elif col == 5:  # Colonna Ora
                    cell.number_format = 'HH:MM'
                elif col == 6:  # Colonna Durata
                    cell.number_format = '0'  # Numero intero
        
        # Nascondi colonne tecniche
        ws.column_dimensions['C'].hidden = True  # Codice insegnamento
        ws.column_dimensions['D'].hidden = True  # Codice CdS
        
        # Aggiungi validazioni per facilitare la compilazione
        if num_rows_to_add > 0:
            last_row = num_rows_to_add + 1
            
            # Validazione Apertura Appelli
            apertura_values = "Sì,No"
            dv_apertura = DataValidation(type="list", formula1=f'"{apertura_values}"')
            dv_apertura.add(f"E2:E{last_row}")
            ws.add_data_validation(dv_apertura)
            
            # Validazione aule
            if aule:
                aule_list = ",".join(aule[:255])  # Limita per evitare errori Excel
                dv_aule = DataValidation(type="list", formula1=f'"{aule_list}"')
                dv_aule.add(f"I2:I{last_row}")
                ws.add_data_validation(dv_aule)
            
            # Validazione verbalizzazione (valori user-friendly)
            verb_values = "Prova finale,Prova finale con pubblicazione,Prova parziale,Prova parziale con pubblicazione"
            dv_verb = DataValidation(type="list", formula1=f'"{verb_values}"')
            dv_verb.add(f"L2:L{last_row}")
            ws.add_data_validation(dv_verb)
            
            # Validazione tipo esame
            tipo_values = "Scritto,Orale,Scritto e orale"
            dv_tipo = DataValidation(type="list", formula1=f'"{tipo_values}"')
            dv_tipo.add(f"M2:M{last_row}")
            ws.add_data_validation(dv_tipo)
        
        # Istruzioni e legenda in foglio separato
        ws_istruzioni = wb.create_sheet("Istruzioni e Legenda")
        istruzioni = [
            "ISTRUZIONI PER LA COMPILAZIONE DEL TEMPLATE ESAMI",
            "",
            "PASSAGGI:",
            "1. Compilare una riga per ogni esame da inserire",
            "2. I campi obbligatori sono evidenziati nella sezione LEGENDA",
            "3. Le validazioni guidano nella compilazione corretta",
            "4. Salvare il file e caricarlo nel sistema",
            "",
            "SUGGERIMENTI GENERALI:",
            "- Gli esami con 'Apertura Appelli = No' non contano per il minimo annuale",
            "- Il formato data è quello italiano (DD-MM-YYYY)",
            "- Se le date di iscrizione sono vuote, verranno calcolate automaticamente",
            "- Per aggiungere più di un appello per lo stesso insegnamento, duplicare la riga e modificare i campi necessari",
            "- Non modificare i valori delle colonne C e D nascoste, sono necessari al sistema per identificare insegnamenti e CdS",
            "",
            "═════════════════════════════════════════════════════════",
            "",
            "LEGENDA CAMPI - VALORI ACCETTATI",
            "",
            "- CDS (Corso di Studio)",
            "   - Precompilato automaticamente",
            "   - Non modificare questo campo",
            "",
            "- INSEGNAMENTO",
            "   - Precompilato automaticamente",
            "   - Non modificare questo campo",
            "",
            "- APERTURA APPELLI (obbligatorio)",
            "   - Sì: l'esame sarà visibile nel calendario pubblico e conta per il numero di esami annuali",
            "   - No: l'esame sarà privato e non conta per il numero di esami annuali",
            "   - Precompilato con 'Sì'",
            "",
            "- DATA (obbligatorio)",
            "   - Formato: DD-MM-YYYY",
            "   - Esempi validi: 15-06-2025, 03-02-2025, 31-12-2024",
            "   - La data deve rispettare le sessioni d'esame configurate",
            "",
            "- ORA (obbligatorio)",
            "   - Formato: HH:MM",
            "   - Esempi validi: 09:00, 14:30, 16:00",
            "   - Usa sempre due cifre per ore e minuti",
            "",
            "- DURATA (obbligatorio)",
            "   - Solo il numero di minuti",
            "   - Esempi: 120 (2 ore), 180 (3 ore), 90 (1,5 ore), solo multipli di 15 minuti",
            "   - Non inserire 'min' o altre unità di misura",
            "",
            "- AULA (opzionale)",
            "   - Selezionare dalla lista a tendina",
            "   - Se vuoto, dovrai assegnare l'aula manualmente dopo l'importazione",
            "   - Studio docente DMI è sempre disponibile",
            "   - Purtroppo non è possibile calcolare in tempo reale la disponibilità delle aule, quindi è necessario verificare manualmente",
            "",
            "- INIZIO ISCRIZIONE (opzionale)",
            "   - Formato: DD-MM-YYYY",
            "   - Se vuoto: calcolato automaticamente (30 giorni prima dell'esame)",
            "   - Esempi: 15-05-2025, 01-01-2025",
            "",
            "- FINE ISCRIZIONE (opzionale)",
            "   - Formato: DD-MM-YYYY",
            "   - Se vuoto: calcolato automaticamente (1 giorno prima dell'esame)",
            "   - Deve essere successiva alla data di inizio iscrizione",
            "",
            "- VERBALIZZAZIONE (precompilato)",
            "   - Prova finale: per esami senza pubblicazione automatica",
            "   - Prova finale con pubblicazione: per esami con pubblicazione del voto",
            "   - Prova parziale: per prove parziali senza pubblicazione",
            "   - Prova parziale con pubblicazione: per prove parziali con pubblicazione del voto",
            "",
            "- TIPO ESAME (opzionale)",
            "   - Scritto: solo prova scritta",
            "   - Orale: solo prova orale",
            "   - Scritto e orale: entrambe le modalità",
            "",
            "- NOTE (opzionale)",
            "   - Campo libero per annotazioni",
            "   - Esempi: 'Portare calcolatrice', 'Prova recupero', 'Sessione straordinaria'",
            "   - Massimo 500 caratteri",
            "",
            "In caso di problemi o domande, contattare il supporto tecnico all'indirizzo opla.dmi.unipg@gmail.com.",
        ]
        
        for row, istruzione in enumerate(istruzioni, 1):
            cell = ws_istruzioni.cell(row, 1, istruzione)
            
            # Formattazione speciale per diversi tipi di contenuto
            if row == 1:  # Titolo principale
                cell.font = Font(bold=True, size=16, name="Arial")
            else:
                cell.font = Font(name="Arial")
            
            cell.alignment = Alignment(horizontal="left", wrap_text=True)
        
        # Imposta larghezza colonna per leggibilità ottimale
        ws_istruzioni.column_dimensions['A'].width = 85
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(buffer, 
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        as_attachment=True, 
                        download_name=f"template_esami_{username}.xlsx")
        
    except Exception as e:
        logging.error(f"Errore template: {e}")
        return jsonify({"error": str(e)}), 500

@import_bp.route('/api/import-exams-from-file', methods=['POST'])
@require_auth
def import_exams_from_file():
    """Importa esami da Excel con mappatura automatica dei valori."""
    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        return jsonify({"success": False, "message": "File XLSX richiesto"}), 400
    
    username = session.get('username')
    anno = request.form.get('anno_accademico')
    bypass = request.form.get('bypass_checks') == 'true'
    
    if not username or not anno:
        return jsonify({"success": False, "message": "Dati mancanti"}), 401
    
    try:
        # Mappature per conversione valori user-friendly -> valori tecnici
        verbalizzazione_map = {
            "Prova finale": "FSS",
            "Prova finale con pubblicazione": "FWP", 
            "Prova parziale": "PAR",
            "Prova parziale con pubblicazione": "PPP"
        }
        
        tipo_esame_map = {
            "Scritto": "S",
            "Orale": "O", 
            "Scritto e orale": "SO"
        }
        
        tipo_appello_map = {
            "Prova finale": "PF",
            "Prova finale con pubblicazione": "PF",
            "Prova parziale": "PP", 
            "Prova parziale con pubblicazione": "PP"
        }
        
        # Processa Excel
        wb = load_workbook(io.BytesIO(file.read()))
        rows = list(wb.active.rows)[1:]  # Salta header
        
        esami = []
        errori = []
        
        for i, row in enumerate(rows, 2):
            values = [cell.value for cell in row]
            
            if not any(values[:2]):  # Salta righe vuote (CdS e Insegnamento)
                continue
                
            try:
                # Estrai valori dalla struttura riorganizzata
                cds_nome = values[0]
                insegnamento_nome = values[1] 
                
                # Valori nascosti per elaborazione (ora subito dopo l'insegnamento)
                codice_insegnamento = values[2] if len(values) > 2 else None
                codice_cds = values[3] if len(values) > 3 else None
                
                apertura_appelli = values[4]
                data = values[5]
                ora = values[6]
                durata = values[7]
                aula = values[8]
                inizio_iscr = values[9]
                fine_iscr = values[10]
                verbalizzazione_friendly = values[11]
                tipo_esame_friendly = values[12]
                note = values[13]
                
                # Controlli minimi per parsing (solo per evitare errori fatali)
                if not all([cds_nome, insegnamento_nome, data, ora]):
                    errori.append(f"Riga {i}: dati obbligatori mancanti")
                    continue
                
                # Parse e validazione data (supporta sia DD-MM-YYYY che YYYY-MM-DD)
                if isinstance(data, (datetime, date)):
                    data_str = data.strftime('%Y-%m-%d')
                else:
                    try:
                        data_str_input = str(data)
                        # Prova prima formato italiano DD-MM-YYYY
                        if len(data_str_input) == 10 and '-' in data_str_input:
                            parts = data_str_input.split('-')
                            if len(parts[0]) == 2:  # Formato DD-MM-YYYY
                                data_obj = datetime.strptime(data_str_input, '%d-%m-%Y')
                            else:  # Formato YYYY-MM-DD
                                data_obj = datetime.strptime(data_str_input, '%Y-%m-%d')
                        else:
                            # Fallback per altri formati
                            data_obj = datetime.strptime(data_str_input, '%d-%m-%Y')
                        data_str = data_obj.strftime('%Y-%m-%d')
                    except:
                        errori.append(f"Riga {i}: formato data non valido (usare DD-MM-YYYY)")
                        continue
                
                # Parse ora
                if isinstance(ora, (time, datetime)):
                    ora_h, ora_m = ora.hour, ora.minute
                elif isinstance(ora, str) and ':' in ora:
                    try:
                        ora_h, ora_m = map(int, ora.split(':'))
                    except:
                        errori.append(f"Riga {i}: formato ora non valido (usare HH:MM)")
                        continue
                else:
                    errori.append(f"Riga {i}: formato ora non valido")
                    continue
                
                # Validazione durata
                if not durata:
                    durata = 120
                else:
                    try:
                        durata = int(durata)
                        if durata <= 0:
                            errori.append(f"Riga {i}: durata deve essere un numero positivo")
                            continue
                    except:
                        errori.append(f"Riga {i}: durata deve essere un numero")
                        continue
                
                # Conversione valori user-friendly
                verbalizzazione = verbalizzazione_map.get(verbalizzazione_friendly, "FSS")
                tipo_esame = tipo_esame_map.get(tipo_esame_friendly, "S") 
                tipo_appello = tipo_appello_map.get(verbalizzazione_friendly, "PF")
                
                # Apertura appelli
                mostra_calendario = str(apertura_appelli).lower() in ['sì', 'si', 'yes', 'true', '1']
                
                # Calcola date iscrizione se mancanti (gestisce formato italiano)
                data_obj = datetime.strptime(data_str, '%Y-%m-%d')
                
                if not inizio_iscr:
                    inizio_iscr = (data_obj - timedelta(days=30)).strftime('%Y-%m-%d')
                elif isinstance(inizio_iscr, (datetime, date)):
                    inizio_iscr = inizio_iscr.strftime('%Y-%m-%d')
                else:
                    try:
                        # Prova formato italiano
                        inizio_obj = datetime.strptime(str(inizio_iscr), '%d-%m-%Y')
                        inizio_iscr = inizio_obj.strftime('%Y-%m-%d')
                    except:
                        inizio_iscr = str(inizio_iscr)
                
                if not fine_iscr:
                    fine_iscr = (data_obj - timedelta(days=1)).strftime('%Y-%m-%d')
                elif isinstance(fine_iscr, (datetime, date)):
                    fine_iscr = fine_iscr.strftime('%Y-%m-%d')
                else:
                    try:
                        # Prova formato italiano
                        fine_obj = datetime.strptime(str(fine_iscr), '%d-%m-%Y')
                        fine_iscr = fine_obj.strftime('%Y-%m-%d')
                    except:
                        fine_iscr = str(fine_iscr)
                
                # Usa codice insegnamento se disponibile, altrimenti cerca per nome
                if not codice_insegnamento:
                    # Fallback: cerca insegnamento per nome (meno affidabile)
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    cursor.execute("SELECT codice FROM insegnamenti WHERE titolo LIKE %s LIMIT 1", 
                                 (f"%{insegnamento_nome}%",))
                    result = cursor.fetchone()
                    cursor.close()
                    release_connection(conn)
                    
                    if not result:
                        errori.append(f"Riga {i}: insegnamento '{insegnamento_nome}' non trovato")
                        continue
                    codice_insegnamento = result[0]
                
                esame = {
                    'insegnamenti': [codice_insegnamento],
                    'docente': username,
                    'anno_accademico': int(anno),
                    'sezioni_appelli': [{
                        'descrizione': f"Appello {insegnamento_nome}",
                        'data_appello': data_str,
                        'ora_h': str(ora_h).zfill(2),
                        'ora_m': str(ora_m).zfill(2),
                        'ora_appello': f"{ora_h:02d}:{ora_m:02d}",
                        'durata': str(durata),
                        'durata_appello': durata,
                        'aula': aula,
                        'inizio_iscrizione': inizio_iscr,
                        'fine_iscrizione': fine_iscr,
                        'verbalizzazione': verbalizzazione,
                        'tipo_esame': tipo_esame,
                        'note_appello': note or "",
                        'tipo_appello': tipo_appello,
                        'mostra_nel_calendario': mostra_calendario,
                        'tipo_iscrizione': "SOC" if tipo_esame == "SO" else tipo_esame,
                        'definizione_appello': 'STD',
                        'gestione_prenotazione': 'STD',
                        'riservato': False,
                        'posti': None,
                        'periodo': 1 if ora_h >= 14 else 0
                    }]
                }
                
                # Controllo vincoli tramite funzione centralizzata
                if not bypass:
                    ok, msg = controlla_vincoli(esame)
                    if not ok:
                        errori.append(f"Riga {i}: {msg}")
                        continue
                
                esami.append(esame)
                
            except Exception as e:
                errori.append(f"Riga {i}: errore elaborazione - {str(e)}")
        
        if not esami:
            return jsonify({
                "success": False,
                "message": f"Nessun esame valido trovato. {len(errori)} errori rilevati.",
                "formatErrors": errori,
                "totalErrors": len(errori)
            }), 400
        
        # Inserisci esami
        successi = 0
        fallimenti = []
        
        for esame in esami:
            try:
                inserisci_esami(esame)
                successi += 1
            except Exception as e:
                # Trova il nome dell'insegnamento per un errore più chiaro
                try:
                    conn_temp = get_db_connection()
                    cursor_temp = conn_temp.cursor()
                    cursor_temp.execute("SELECT titolo FROM insegnamenti WHERE codice = %s", (esame['insegnamenti'][0],))
                    result = cursor_temp.fetchone()
                    titolo = result[0] if result else esame['insegnamenti'][0]
                    cursor_temp.close()
                    release_connection(conn_temp)
                except:
                    titolo = esame['insegnamenti'][0]
                
                fallimenti.append(f"Inserimento {titolo}: {str(e)}")
        
        # Prepara risposta con errori categorizzati
        message = f"{successi} esami inseriti con successo"
        if bypass:
            message += " (controlli bypassati)"
        
        total_errors = len(errori) + len(fallimenti)
        if total_errors > 0:
            message += f", {total_errors} errori rilevati"
        
        return jsonify({
            "success": successi > 0,
            "message": message,
            "importedCount": successi,
            "totalErrors": total_errors,
            "formatErrors": errori,
            "insertionErrors": fallimenti
        })
    
    except Exception as e:
        logging.error(f"Errore import: {e}")
        return jsonify({"success": False, "message": f"Errore durante l'importazione: {e}"}), 500