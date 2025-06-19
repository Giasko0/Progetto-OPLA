from flask import Blueprint, request, jsonify, Response, session
from db import get_db_connection, release_connection
from psycopg2.extras import DictCursor
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging

controllo_esami_minimi_bp = Blueprint('controllo_esami_minimi', __name__, url_prefix='/api/oh-issa')

@controllo_esami_minimi_bp.route('/get-docenti-by-anno')
def get_docenti_by_anno():
    """Ottiene tutti i docenti che insegnano in un anno accademico"""
    if not session.get('permessi_admin'):
        return jsonify({'status': 'error', 'message': 'Accesso non autorizzato'}), 401
    
    anno = request.args.get('anno')
    if not anno:
        return jsonify({'error': 'Anno accademico non specificato'}), 400
    
    try:
        anno = int(anno)
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=DictCursor)
        
        cursor.execute("""
            SELECT DISTINCT u.username, u.nome, u.cognome, u.matricola
            FROM utenti u
            JOIN insegnamento_docente id ON u.username = id.docente
            WHERE id.annoaccademico = %s
            ORDER BY u.cognome, u.nome
        """, (anno,))
        
        docenti = []
        for row in cursor.fetchall():
            docenti.append({
                'username': row['username'],
                'nome': row['nome'],
                'cognome': row['cognome'],
                'matricola': row['matricola']
            })
        
        return jsonify(docenti)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            release_connection(conn)

@controllo_esami_minimi_bp.route('/controlla-esami-minimi')
def controlla_esami_minimi():
    """Controlla il numero di esami per ogni insegnamento"""
    if not session.get('permessi_admin'):
        return jsonify({'status': 'error', 'message': 'Accesso non autorizzato'}), 401
    
    anno = request.args.get('anno')
    cds_filter = request.args.get('cds')
    docente_filter = request.args.get('docente')
    
    if not anno:
        return jsonify({'error': 'Anno accademico non specificato'}), 400
    
    try:
        anno = int(anno)
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=DictCursor)
        
        # Query base per ottenere tutti gli insegnamenti dell'anno
        base_query = """
            SELECT 
                i.id as insegnamento_id,
                i.codice as insegnamento_codice,
                i.titolo as insegnamento_titolo,
                ic.cds as cds_codice,
                c.nome_corso as cds_nome,
                ic.curriculum_codice,
                ic.anno_corso,
                ic.semestre,
                COUNT(CASE WHEN e.mostra_nel_calendario = TRUE AND e.tipo_appello != 'PP' THEN e.id END) as numero_esami,
                u.username as docente_username,
                u.nome as docente_nome,
                u.cognome as docente_cognome
            FROM insegnamenti i
            JOIN insegnamenti_cds ic ON i.id = ic.insegnamento
            JOIN cds c ON ic.cds = c.codice AND ic.anno_accademico = c.anno_accademico AND ic.curriculum_codice = c.curriculum_codice
            LEFT JOIN insegnamento_docente id ON i.id = id.insegnamento AND id.annoaccademico = %s
            LEFT JOIN utenti u ON id.docente = u.username
            LEFT JOIN esami e ON i.id = e.insegnamento AND e.anno_accademico = %s AND e.cds = ic.cds AND e.curriculum_codice = ic.curriculum_codice
            WHERE ic.anno_accademico = %s
        """
        
        params = [anno, anno, anno]
        
        # Aggiungi filtri se specificati
        if cds_filter:
            base_query += " AND ic.cds = %s"
            params.append(cds_filter)
        
        if docente_filter:
            base_query += " AND u.username = %s"
            params.append(docente_filter)
        
        base_query += """
            GROUP BY i.id, i.codice, i.titolo, ic.cds, c.nome_corso, ic.curriculum_codice, 
                     ic.anno_corso, ic.semestre, u.username, u.nome, u.cognome
            ORDER BY ic.cds, i.titolo, u.cognome, u.nome
        """
        
        cursor.execute(base_query, params)
        rows = cursor.fetchall()
        
        # Raggruppa i risultati per insegnamento
        insegnamenti_dict = {}
        
        for row in rows:
            ins_key = f"{row['insegnamento_id']}_{row['cds_codice']}_{row['curriculum_codice']}"
            
            if ins_key not in insegnamenti_dict:
                insegnamenti_dict[ins_key] = {
                    'id': row['insegnamento_id'],
                    'codice': row['insegnamento_codice'],
                    'titolo': row['insegnamento_titolo'],
                    'cds_codice': row['cds_codice'],
                    'cds_nome': row['cds_nome'],
                    'curriculum_codice': row['curriculum_codice'],
                    'anno_corso': row['anno_corso'],
                    'semestre': row['semestre'],
                    'numero_esami': row['numero_esami'],
                    'docenti': []
                }
            
            # Aggiungi il docente se presente e non già aggiunto
            if row['docente_username']:
                docente = {
                    'username': row['docente_username'],
                    'nome': row['docente_nome'],
                    'cognome': row['docente_cognome']
                }
                
                # Verifica che il docente non sia già presente
                if not any(d['username'] == docente['username'] for d in insegnamenti_dict[ins_key]['docenti']):
                    insegnamenti_dict[ins_key]['docenti'].append(docente)
        
        # Converte in lista
        insegnamenti = list(insegnamenti_dict.values())
        
        # Statistiche
        total = len(insegnamenti)
        conformi = len([ins for ins in insegnamenti if ins['numero_esami'] >= 8])
        non_conformi = total - conformi
        
        result = {
            'anno_accademico': anno,
            'cds_filter': cds_filter,
            'docente_filter': docente_filter,
            'statistiche': {
                'totale_insegnamenti': total,
                'conformi': conformi,
                'non_conformi': non_conformi,
                'percentuale_conformi': round(conformi / total * 100, 1) if total > 0 else 0
            },
            'insegnamenti': insegnamenti
        }
        
        return jsonify(result)
        
    except Exception as e:
        logging.error(f"Errore nel controllo esami minimi: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            release_connection(conn)

@controllo_esami_minimi_bp.route('/esporta-controllo-esami-minimi')
def esporta_controllo_esami_minimi():
    """Esporta il report di controllo esami minimi in formato Excel"""
    if not session.get('permessi_admin'):
        return jsonify({'status': 'error', 'message': 'Accesso non autorizzato'}), 401
    
    anno = request.args.get('anno')
    cds_filter = request.args.get('cds')
    docente_filter = request.args.get('docente')
    
    if not anno:
        return jsonify({'error': 'Anno accademico non specificato'}), 400
    
    try:
        # Ottieni i dati utilizzando la stessa logica dell'endpoint di controllo
        params = {'anno': anno}
        if cds_filter:
            params['cds'] = cds_filter
        if docente_filter:
            params['docente'] = docente_filter
        
        # Simula la chiamata interna
        from flask import g
        g.simulate_request = True
        
        with controllo_esami_minimi_bp.test_request_context(query_string=params):
            # Richiama la funzione di controllo direttamente
            response_data = controlla_esami_minimi()
            if isinstance(response_data, tuple):
                data = response_data[0].get_json()
            else:
                data = response_data.get_json()
        
        if 'error' in data:
            return jsonify({'error': data['error']}), 500
        
        # Crea il workbook Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"Controllo Esami Minimi {anno}"
        
        # Stili
        header_font = Font(bold=True, size=12, name="Arial")
        title_font = Font(bold=True, size=14, name="Arial")
        normal_font = Font(size=10, name="Arial")
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        conforme_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        non_conforme_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        
        # Titolo del report
        current_row = 1
        title_cell = sheet.cell(row=current_row, column=1, value=f"CONTROLLO ESAMI MINIMI - A.A. {anno}/{anno+1}")
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        sheet.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2
        
        # Informazioni filtri
        if cds_filter or docente_filter:
            filter_text = "Filtri applicati: "
            if cds_filter:
                filter_text += f"CdS: {cds_filter} "
            if docente_filter:
                filter_text += f"Docente: {docente_filter}"
            
            filter_cell = sheet.cell(row=current_row, column=1, value=filter_text)
            filter_cell.font = normal_font
            sheet.merge_cells(f'A{current_row}:G{current_row}')
            current_row += 2
        
        # Statistiche riassuntive
        stats = data['statistiche']
        stats_data = [
            ["Totale Insegnamenti:", stats['totale_insegnamenti']],
            ["Conformi (≥8 esami):", stats['conformi']],
            ["Non Conformi (<8 esami):", stats['non_conformi']],
            ["Percentuale Conformi:", f"{stats['percentuale_conformi']}%"]
        ]
        
        for stat_label, stat_value in stats_data:
            sheet.cell(row=current_row, column=1, value=stat_label).font = header_font
            sheet.cell(row=current_row, column=2, value=stat_value).font = normal_font
            current_row += 1
        
        current_row += 2
        
        # Header della tabella
        headers = ["CdS", "Insegnamento", "Codice", "Docenti", "Anno", "Sem.", "Esami", "Status"]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=current_row, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border
        
        current_row += 1
        
        # Raggruppa per CdS
        insegnamenti_per_cds = {}
        for ins in data['insegnamenti']:
            cds_key = f"{ins['cds_codice']} - {ins['cds_nome']}"
            if cds_key not in insegnamenti_per_cds:
                insegnamenti_per_cds[cds_key] = []
            insegnamenti_per_cds[cds_key].append(ins)
        
        # Popola i dati
        for cds_name in sorted(insegnamenti_per_cds.keys()):
            insegnamenti = insegnamenti_per_cds[cds_name]
            
            # Ordina per numero esami (crescente) e poi per titolo
            insegnamenti.sort(key=lambda x: (x['numero_esami'], x['titolo']))
            
            for ins in insegnamenti:
                docenti_text = ", ".join([f"{d['cognome']} {d['nome']}" for d in ins['docenti']]) if ins['docenti'] else "Nessun docente"
                is_conforme = ins['numero_esami'] >= 8
                status_text = "✓ Conforme" if is_conforme else "⚠ Non conforme"
                
                row_data = [
                    ins['cds_codice'],
                    ins['titolo'],
                    ins['codice'],
                    docenti_text,
                    ins['anno_corso'],
                    ins['semestre'],
                    ins['numero_esami'],
                    status_text
                ]
                
                # Applica il colore di sfondo in base alla conformità
                fill_color = conforme_fill if is_conforme else non_conforme_fill
                
                for col_num, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=current_row, column=col_num, value=value)
                    cell.font = normal_font
                    cell.border = thin_border
                    cell.fill = fill_color
                    
                    if col_num in [1, 5, 6, 7]:  # CdS, Anno, Semestre, Esami - centra
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = left_alignment
                
                current_row += 1
        
        # Imposta larghezza colonne
        column_widths = [10, 50, 15, 40, 8, 8, 10, 15]
        for col_num, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(col_num)].width = width
        
        # Genera il file
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Nome file
        filename = f"controllo_esami_minimi_{anno}"
        if cds_filter:
            filename += f"_{cds_filter}"
        if docente_filter:
            filename += f"_{docente_filter}"
        filename += ".xlsx"
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
        
    except Exception as e:
        logging.error(f"Errore durante l'esportazione controllo esami: {str(e)}", exc_info=True)
        return jsonify({"error": f"Errore interno del server: {str(e)}"}), 500
