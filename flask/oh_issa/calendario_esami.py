from flask import Blueprint, request, jsonify, Response
from db import get_db_connection, release_connection
from psycopg2.extras import DictCursor
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime, timedelta
import logging
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

calendario_esami_bp = Blueprint('calendario_esami', __name__, url_prefix='/api/oh-issa')

def filtra_date_14_giorni(date_list):
    """
    Filtra una lista di date mantenendo solo quelle a distanza minima di 14 giorni.
    Prende la prima, poi ignora quelle entro 14 giorni, seleziona la più vicina >= 14 giorni.
    """
    if not date_list:
        return []
    
    # Ordina le date
    date_ordinate = sorted(date_list)
    date_filtrate = [date_ordinate[0]]
    
    for data in date_ordinate[1:]:
        ultima_selezionata = date_filtrate[-1]
        if (data - ultima_selezionata).days >= 14:
            date_filtrate.append(data)
    
    return date_filtrate

def _get_esami_per_insegnamento(cursor, cds_code, anno_accademico, curriculum, insegnamenti_base, sessioni_calendario):
    """
    Helper function per ottenere tutti gli esami (ufficiali e non ufficiali) per ogni insegnamento.
    Restituisce un dizionario {insegnamento_id: [{'data_appello': date, 'mostra_nel_calendario': bool}, ...]}
    """
    # Ottieni esami diretti
    cursor.execute("""
        SELECT e.insegnamento, e.data_appello, e.mostra_nel_calendario
        FROM esami e
        JOIN insegnamenti_cds ic ON e.insegnamento = ic.insegnamento
        WHERE e.cds = %s AND e.anno_accademico = %s
          AND e.curriculum_codice IN (%s, 'GEN')
          AND e.data_appello >= %s
          AND e.mostra_nel_calendario = TRUE
          AND ic.cds = %s AND ic.anno_accademico = %s
          AND (ic.curriculum_codice = %s OR ic.curriculum_codice = 'GEN')
          AND (ic.inserire_esami = TRUE OR ic.master IS NULL)
        ORDER BY e.data_appello
    """, (cds_code, anno_accademico, curriculum, f"{anno_accademico}-01-01",
          cds_code, anno_accademico, curriculum))
    
    esami_diretti = cursor.fetchall()
    
    # Ottieni esami dal master
    cursor.execute("""
        SELECT ic.insegnamento, e.data_appello, e.mostra_nel_calendario
        FROM insegnamenti_cds ic
        JOIN esami e ON ic.master = e.insegnamento
        WHERE ic.cds = %s AND ic.anno_accademico = %s
          AND (ic.curriculum_codice = %s OR ic.curriculum_codice = 'GEN')
          AND ic.inserire_esami = FALSE
          AND ic.master IS NOT NULL
          AND e.anno_accademico = %s
          AND e.mostra_nel_calendario = TRUE
        ORDER BY e.data_appello
    """, (cds_code, anno_accademico, curriculum, anno_accademico))
    
    esami_master = cursor.fetchall()
    
    # Combina esami in un dizionario per insegnamento
    esami_per_insegnamento = {}
    
    for esame in esami_diretti:
        if esame['insegnamento'] not in esami_per_insegnamento:
            esami_per_insegnamento[esame['insegnamento']] = []
        esami_per_insegnamento[esame['insegnamento']].append({
            'data_appello': esame['data_appello'],
            'mostra_nel_calendario': esame['mostra_nel_calendario']
        })
    
    for esame in esami_master:
        if esame['insegnamento'] not in esami_per_insegnamento:
            esami_per_insegnamento[esame['insegnamento']] = []
        esami_per_insegnamento[esame['insegnamento']].append({
            'data_appello': esame['data_appello'],
            'mostra_nel_calendario': esame['mostra_nel_calendario']
        })
    
    # Aggiungi esami non ufficiali SOLO per insegnamenti annuali nella sessione anticipata
    if sessioni_calendario['anticipata']['inizio']:
        for insegnamento in insegnamenti_base:
            # Verifica che sia un insegnamento annuale (semestre = 3)
            if insegnamento['semestre'] == 3:
                # Recupera tutte le date già presenti per questo insegnamento
                date_esistenti = {e['data_appello'] for e in esami_per_insegnamento.get(insegnamento['id'], [])}
                
                cursor.execute("""
                    SELECT e.data_appello
                    FROM esami e
                    WHERE e.insegnamento = %s
                      AND e.cds = %s
                      AND e.anno_accademico = %s
                      AND (e.curriculum_codice = %s OR e.curriculum_codice = 'GEN')
                      AND e.data_appello >= %s
                      AND e.data_appello BETWEEN %s AND %s
                """, (
                    insegnamento['id'],
                    cds_code,
                    anno_accademico,
                    insegnamento['curriculum_codice'],
                    f"{anno_accademico}-01-01",
                    sessioni_calendario['anticipata']['inizio'],
                    sessioni_calendario['anticipata']['fine']
                ))
                
                esami_non_ufficiali_raw = [row['data_appello'] for row in cursor.fetchall()]
                
                # Filtra le date già esistenti (già recuperate come ufficiali)
                esami_non_ufficiali_raw = [data for data in esami_non_ufficiali_raw if data not in date_esistenti]
                
                # Filtra a 14 giorni di distanza
                esami_non_ufficiali_filtrati = filtra_date_14_giorni(esami_non_ufficiali_raw)
                
                # Inizializza la lista se non esiste
                if insegnamento['id'] not in esami_per_insegnamento:
                    esami_per_insegnamento[insegnamento['id']] = []
                
                # Aggiungi solo gli esami filtrati come non ufficiali
                for data in esami_non_ufficiali_filtrati:
                    esami_per_insegnamento[insegnamento['id']].append({
                        'data_appello': data,
                        'mostra_nel_calendario': False
                    })
    
    return esami_per_insegnamento

# API per ottenere i curriculum per un corso di studi e anno
@calendario_esami_bp.route('/get-curriculum-by-cds')
def get_curriculum_by_cds():
    try:
        cds_code = request.args.get('cds')
        anno_accademico = request.args.get('anno')
        
        if not cds_code or not anno_accademico:
            return jsonify({'error': 'Parametri mancanti'})
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=DictCursor)
        
        cursor.execute("""
            SELECT DISTINCT curriculum_codice, curriculum_nome
            FROM cds
            WHERE codice = %s AND anno_accademico = %s
            ORDER BY curriculum_codice
        """, (cds_code, anno_accademico))
        
        curriculum_list = [{'codice': row['curriculum_codice'], 'nome': row['curriculum_nome']} for row in cursor.fetchall()]
        
        return jsonify(curriculum_list)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            release_connection(conn)

# API per ottenere i dati del calendario esami
@calendario_esami_bp.route('/get-calendario-esami')
def get_calendario_esami():
  try:
    cds_code = request.args.get('cds')
    anno_accademico = request.args.get('anno')
    curriculum = request.args.get('curriculum')
    
    if not cds_code or not anno_accademico or not curriculum:
      return jsonify({'error': 'Parametri mancanti'})
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=DictCursor)
    
    # Ottieni i dettagli del corso di studi
    cursor.execute("""
      SELECT nome_corso, curriculum_codice, curriculum_nome
      FROM cds
      WHERE codice = %s AND anno_accademico = %s AND curriculum_codice = %s
    """, (cds_code, anno_accademico, curriculum))
    
    cds_info = cursor.fetchone()
    if not cds_info:
      return jsonify({'error': 'Corso di studi non trovato'})
    
    # Ottieni tutti gli insegnamenti e le loro informazioni
    cursor.execute("""
      SELECT i.id, i.codice, i.titolo, ic.anno_corso, ic.semestre, 
             ic.curriculum_codice, ic.inserire_esami, ic.master
      FROM insegnamenti i
      JOIN insegnamenti_cds ic ON i.id = ic.insegnamento
      WHERE ic.cds = %s AND ic.anno_accademico = %s 
        AND (ic.curriculum_codice = %s OR ic.curriculum_codice = 'GEN')
      ORDER BY ic.anno_corso, i.titolo
    """, (cds_code, anno_accademico, curriculum))
    
    insegnamenti_base = cursor.fetchall()
    
    # Query per ottenere le sessioni
    cursor.execute("""
      SELECT tipo_sessione, inizio, fine
      FROM sessioni
      WHERE cds = %s AND anno_accademico = %s 
        AND (curriculum_codice = %s OR curriculum_codice = 'GEN')
      ORDER BY inizio
    """, (cds_code, anno_accademico, curriculum))
    
    sessioni = cursor.fetchall()
    
    # Prepara le sessioni per il calendario
    sessioni_calendario = {
      'anticipata': {'nome': 'Sessione Anticipata', 'inizio': None, 'fine': None},
      'estiva': {'nome': 'Sessione Estiva', 'inizio': None, 'fine': None},
      'autunnale': {'nome': 'Sessione Autunnale', 'inizio': None, 'fine': None},
      'invernale': {'nome': 'Sessione Invernale', 'inizio': None, 'fine': None}
    }
    
    # Popola le date delle sessioni
    for sessione in sessioni:
      tipo_sessione, inizio, fine = sessione
      if tipo_sessione in sessioni_calendario:
        sessioni_calendario[tipo_sessione]['inizio'] = inizio
        sessioni_calendario[tipo_sessione]['fine'] = fine
    
    # Se non c'è sessione anticipata, cerca quella invernale dell'anno precedente
    if not sessioni_calendario['anticipata']['inizio']:
      cursor.execute("""
        SELECT inizio, fine
        FROM sessioni
        WHERE cds = %s AND anno_accademico = %s AND tipo_sessione = 'invernale'
          AND (curriculum_codice = %s OR curriculum_codice = 'GEN')
      """, (cds_code, int(anno_accademico) - 1, curriculum))
      
      sessione_precedente = cursor.fetchone()
      if sessione_precedente:
        sessioni_calendario['anticipata']['inizio'] = sessione_precedente[0]
        sessioni_calendario['anticipata']['fine'] = sessione_precedente[1]
    
    # Usa la funzione helper per ottenere gli esami
    esami_per_insegnamento = _get_esami_per_insegnamento(
        cursor, cds_code, anno_accademico, curriculum, insegnamenti_base, sessioni_calendario
    )
    
    # Costruisci la lista finale degli insegnamenti
    insegnamenti = []
    for row in insegnamenti_base:
      insegnamento_id = row['id']
      # Restituisci tutti gli esami con il flag mostra_nel_calendario
      esami_completi = [
        {
          'data_appello': e['data_appello'],
          'mostra_nel_calendario': e['mostra_nel_calendario']
        }
        for e in esami_per_insegnamento.get(insegnamento_id, [])
      ]
      
      insegnamenti.append({
        'id': insegnamento_id,
        'codice': row['codice'],
        'titolo': row['titolo'],
        'anno_corso': row['anno_corso'],
        'semestre': row['semestre'],
        'curriculum_codice': row['curriculum_codice'],
        'esami': esami_completi
      })
    
    # Ordina insegnamenti per semestre (3, 1, 2) e poi alfabeticamente
    semestre_order = {3: 0, 1: 1, 2: 2}
    insegnamenti.sort(key=lambda x: (semestre_order.get(x['semestre'], 999), x['titolo']))
    
    return jsonify({
      'nome_corso': cds_info['nome_corso'],
      'curriculum': cds_info['curriculum_nome'],
      'insegnamenti': insegnamenti,
      'sessioni': sessioni_calendario
    })
  except Exception as e:
    return jsonify({"error": str(e)}), 500
  finally:
    if 'cursor' in locals() and cursor:
      cursor.close()
    if 'conn' in locals() and conn:
      release_connection(conn)

# API per esportare il calendario esami in formato XLSX
@calendario_esami_bp.route('/esporta-calendario-esami')
def esporta_calendario_esami():
    try:
        cds_code = request.args.get('cds')
        anno_accademico_str = request.args.get('anno')
        curriculum = request.args.get('curriculum')
        
        if not all([cds_code, anno_accademico_str, curriculum]):
            return jsonify({'error': 'Parametri mancanti'}), 400
        
        try:
            anno_accademico = int(anno_accademico_str)
        except ValueError:
            return jsonify({'error': 'Anno accademico non valido'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=DictCursor)
        
        # Info corso
        cursor.execute("""
            SELECT nome_corso, curriculum_codice, curriculum_nome 
            FROM cds
            WHERE codice = %s AND anno_accademico = %s AND curriculum_codice = %s
        """, (cds_code, anno_accademico, curriculum))
        
        cds_info = cursor.fetchone()
        if not cds_info:
            return jsonify({'error': 'Corso di studi non trovato'}), 404
        
        # Ottieni insegnamenti base
        cursor.execute("""
            SELECT i.id, i.codice, i.titolo, ic.anno_corso, ic.semestre, 
                   ic.curriculum_codice, ic.inserire_esami, ic.master
            FROM insegnamenti i
            JOIN insegnamenti_cds ic ON i.id = ic.insegnamento
            WHERE ic.cds = %s AND ic.anno_accademico = %s 
              AND (ic.curriculum_codice = %s OR ic.curriculum_codice = 'GEN')
            ORDER BY ic.anno_corso, i.titolo
        """, (cds_code, anno_accademico, curriculum))
        
        insegnamenti_base = cursor.fetchall()

        # Query per le sessioni
        cursor.execute("""
            SELECT tipo_sessione, inizio, fine
            FROM sessioni
            WHERE cds = %s AND anno_accademico = %s 
              AND (curriculum_codice = %s OR curriculum_codice = 'GEN')
            ORDER BY inizio
        """, (cds_code, anno_accademico, curriculum))
        
        sessioni = cursor.fetchall()
        
        sessioni_dict = {s['tipo_sessione']: {'inizio': s['inizio'], 'fine': s['fine']} for s in sessioni}
        
        sessioni_calendario = {
            'anticipata': {'nome': 'Sessione Anticipata', 'inizio': None, 'fine': None},
            'estiva': {'nome': 'Sessione Estiva', 'inizio': None, 'fine': None},
            'autunnale': {'nome': 'Sessione Autunnale', 'inizio': None, 'fine': None},
            'invernale': {'nome': 'Sessione Invernale', 'inizio': None, 'fine': None}
        }
        
        for tipo_sessione, data_sessione in sessioni_dict.items():
            if tipo_sessione in sessioni_calendario:
                sessioni_calendario[tipo_sessione].update(data_sessione)
        
        # Usa la funzione helper per ottenere gli esami
        esami_per_insegnamento = _get_esami_per_insegnamento(
            cursor, cds_code, anno_accademico, curriculum, insegnamenti_base, sessioni_calendario
        )
        
        # Costruisci lista insegnamenti finale
        insegnamenti = []
        for row in insegnamenti_base:
            insegnamenti.append({
                'id': row['id'],
                'codice': row['codice'],
                'titolo': row['titolo'],
                'anno_corso': row['anno_corso'],
                'semestre': row['semestre'],
                'curriculum_codice': row['curriculum_codice'],
                'esami': esami_per_insegnamento.get(row['id'], [])
            })
        
        # Ordina insegnamenti per semestre (3, 1, 2) e poi alfabeticamente
        semestre_order = {3: 0, 1: 1, 2: 2}
        insegnamenti.sort(key=lambda x: (semestre_order.get(x['semestre'], 999), x['titolo']))
        
        # Raggruppa gli insegnamenti per anno di corso
        insegnamenti_per_anno = {}
        for ins in insegnamenti:
            anno = ins['anno_corso'] or 1
            insegnamenti_per_anno.setdefault(anno, []).append(ins)
        
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"Calendario {cds_code} {anno_accademico}"
        
        header_font = Font(bold=True, size=14, name="Arial")
        year_header_font = Font(bold=True, size=16, color="FFFFFF", name="Arial")
        insegnamento_font = Font(bold=False, size=12, name="Arial")
        blue_font = Font(bold=False, size=12, color="00BFFF", name="Arial")
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        center_wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        year_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        # Sfondo bianco per insegnamenti e date
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        current_row = 1
        ordine_sessioni = ['anticipata', 'estiva', 'autunnale', 'invernale']
        
        for anno_corso in sorted(insegnamenti_per_anno.keys()):
            if not insegnamenti_per_anno[anno_corso]:
                continue
            
            num_colonne_sessioni = sum(1 for tipo in ordine_sessioni 
                                     if tipo in sessioni_dict and sessioni_calendario[tipo]['inizio'])
            total_columns = 1 + num_colonne_sessioni
            
            year_cell = sheet.cell(row=current_row, column=1, value=f"{anno_corso}° Anno")
            year_cell.font = year_header_font
            year_cell.alignment = center_alignment
            year_cell.fill = year_fill
            year_cell.border = thin_border
            
            end_col = min(total_columns, 26)
            sheet.merge_cells(f'A{current_row}:{get_column_letter(end_col)}{current_row}')
            current_row += 1
            
            insegnamento_cell = sheet.cell(row=current_row, column=1, value='INSEGNAMENTO')
            insegnamento_cell.font = header_font
            insegnamento_cell.alignment = center_alignment
            insegnamento_cell.fill = header_fill
            insegnamento_cell.border = thin_border
            
            current_col = 2
            for tipo_sessione in ordine_sessioni:
                if tipo_sessione in sessioni_dict and sessioni_calendario[tipo_sessione]['inizio']:
                    session_cell = sheet.cell(row=current_row, column=current_col, 
                                            value=sessioni_calendario[tipo_sessione]['nome'].upper())
                    session_cell.font = header_font
                    session_cell.alignment = center_alignment
                    session_cell.fill = header_fill
                    session_cell.border = thin_border
                    current_col += 1
            
            current_row += 1
            
            col_positions = [1]
            data_col = 2
            for tipo_sessione in ordine_sessioni:
                if tipo_sessione in sessioni_dict and sessioni_calendario[tipo_sessione]['inizio']:
                    col_positions.append(data_col)
                    data_col += 1
            
            for insegnamento in insegnamenti_per_anno[anno_corso]:
                semestre = insegnamento.get('semestre')
                if semestre == 1:
                    semestre_str = "Primo semestre"
                elif semestre == 2:
                    semestre_str = "Secondo semestre"
                elif semestre == 3:
                    semestre_str = "Annuale"
                else:
                    semestre_str = ""
                
                name_cell = sheet.cell(row=current_row, column=1)
                
                if semestre_str:
                    # Crea RichText con titolo e semestre in stili diversi
                    rt = CellRichText(
                        TextBlock(InlineFont(sz=12, color="000000", rFont="Arial"), insegnamento['titolo']),
                        TextBlock(InlineFont(sz=10, color="808080", rFont="Arial"), f"\n{semestre_str}")
                    )
                    name_cell.value = rt
                else:
                    name_cell.value = insegnamento['titolo']
                
                name_cell.alignment = left_wrap_alignment
                name_cell.border = thin_border
                name_cell.fill = white_fill
                col_index = 0
                for tipo_sessione in ordine_sessioni:
                    if tipo_sessione in sessioni_dict and sessioni_calendario[tipo_sessione]['inizio']:
                        col_index += 1
                        # Filtra esami per sessione corrente
                        esami_sessione = [
                            esame for esame in insegnamento['esami']
                            if sessioni_calendario[tipo_sessione]['inizio'] <= esame['data_appello'] <= sessioni_calendario[tipo_sessione]['fine']
                        ]
                        
                        cell_value = ""
                        cell_font = insegnamento_font
                        
                        if esami_sessione:
                            # Separa esami ufficiali e non ufficiali
                            esami_ufficiali = [e for e in esami_sessione if e['mostra_nel_calendario']]
                            esami_non_ufficiali = [e for e in esami_sessione if not e['mostra_nel_calendario']]
                            
                            # Tutti gli esami annuali dell'anticipata in blu
                            if semestre == 3 and tipo_sessione == 'anticipata':
                                tutte_date = sorted(set([e['data_appello'] for e in esami_sessione]))
                                date_list = [d.strftime('%d/%m/%Y') for d in tutte_date]
                                cell_value = '\n'.join(date_list)
                                cell_font = blue_font
                            else:
                                # Per altre sessioni mostra solo esami ufficiali
                                if esami_ufficiali:
                                    esami_ufficiali.sort(key=lambda x: x['data_appello'])
                                    date_list = [esame['data_appello'].strftime('%d/%m/%Y') for esame in esami_ufficiali]
                                    cell_value = '\n'.join(date_list)
                        else:
                            cell_value = "--"
                        
                        if col_index < len(col_positions):
                            date_cell = sheet.cell(row=current_row, column=col_positions[col_index], value=cell_value)
                            date_cell.alignment = center_wrap_alignment
                            date_cell.border = thin_border
                            date_cell.fill = white_fill
                            date_cell.font = cell_font
                
                current_row += 1
            
            current_row += 1
        
        sheet.column_dimensions['A'].width = 60
        for col_num in range(2, data_col):
            sheet.column_dimensions[get_column_letter(col_num)].width = 40
        
        # Imposta altezza righe: header 45, insegnamenti 60
        for row_num in range(1, current_row):
            row_obj = sheet.row_dimensions[row_num]
            # Verifica se è un header (anno o intestazione colonne)
            cell_value = sheet.cell(row=row_num, column=1).value
            if cell_value and (str(cell_value).endswith('° Anno') or cell_value == 'INSEGNAMENTO'):
                row_obj.height = 45
            else:
                row_obj.height = 60
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        file_size = len(output.getvalue())
        if file_size == 0:
            return jsonify({'error': 'Errore nella generazione del file Excel'}), 500
        
        filename = f"calendario_esami_{cds_code}_{anno_accademico_str}_{curriculum.replace(' ', '_').replace('/', '_')}.xlsx"
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Length': str(file_size)
            }
        )
        
    except Exception as e:
        logging.error(f"Errore durante l'esportazione XLSX: {str(e)}", exc_info=True)
        return jsonify({"error": f"Errore interno del server: {str(e)}"}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            release_connection(conn)